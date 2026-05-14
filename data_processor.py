#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
就业数据自动化处理工具
========================
自动完成Excel数据表格的清洗、验证、合并与标注。

处理流程：
  第一阶段：加载书院表格，解析企业信息TXT，匹配并修正企业信息
  第二阶段：填充全校已有人数列、校内人数配额检查、注册资本审核、个体工商户审核
  第三阶段：将审核通过的数据合并到全校总表
  第四阶段：输出处理后的书院表格、更新后总表、出错学生信息

用法：
  python data_processor.py --new "书院待筛选表格.xlsx" --total "全校数据总表格.xlsx" --enterprise "所有企业信息.txt"
"""

import os
import sys
import re
import argparse
import logging
from datetime import datetime
from copy import copy

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ============================================================
# 常量定义
# ============================================================
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
RED_FILL = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
HEADER_FONT = Font(bold=True)

UNIT_NAME_COL = "单位"
CREDIT_CODE_COL = "社会信用代码"
REG_CAPITAL_COL = "注册资本"
REASON_COL = "原因"
STATUS_COL = "状态"

STUDENT_NAME_PATTERNS = ["学生姓名", "姓名", "学生名称"]

MAX_PEOPLE_PER_UNIT = 3
THIS_SUBMIT_COL = "本次上报人数"
TOTAL_COUNT_COL = "全校已有人数"


# ============================================================
# 日志配置
# ============================================================
def setup_logging():
    logging.basicConfig(
        level=logging.INFO,
        format="[%(asctime)s] %(message)s",
        datefmt="%H:%M:%S"
    )
    return logging.getLogger(__name__)


logger = setup_logging()

# 操作日志：记录本次处理对表格做过的所有操作，最后输出
operations_log = []


# ============================================================
# 命令行参数解析
# ============================================================
def parse_args():
    parser = argparse.ArgumentParser(
        description="就业数据自动化处理工具",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例用法：
  python data_processor.py
  python data_processor.py --new "input/书院待筛选表格.xlsx" --total "input/全校数据总表格.xlsx" --enterprise "input/所有企业信息.txt"
        """
    )
    parser.add_argument("--new", type=str, default="input/书院待筛选表格.xlsx",
                        help="书院待筛选表格文件路径（默认：input/书院待筛选表格.xlsx）")
    parser.add_argument("--total", type=str, default="input/全校数据总表格.xlsx",
                        help="全校数据总表格文件路径（默认：input/全校数据总表格.xlsx）")
    parser.add_argument("--enterprise", type=str, default="input/所有企业信息.txt",
                        help="所有企业信息文件路径（默认：input/所有企业信息.txt）")
    return parser.parse_args()


# ============================================================
# 企业信息解析
# ============================================================
def parse_registered_capital(value):
    """解析注册资本字符串，返回数值（单位：万元）"""
    if value is None or str(value).strip() == "":
        return 0.0

    s = str(value).strip()
    # 去除常见中文单位
    s = s.replace("万元", "").replace("万", "").replace("元", "").replace(" ", "")
    # 去除逗号千位分隔符
    s = s.replace(",", "").replace("，", "")

    try:
        return float(s)
    except ValueError:
        return 0.0


# ============================================================
# 智能企业信息解析（支持任意格式）
# ============================================================

# ---- 正则模式定义 ----
# 单位名称匹配：先找标签格式，再找后缀格式
RE_COMPANY_LABEL = re.compile(
    r'(?:单位名称|企业名称|公司名称|名称)[：:\s]*([^\n,，\t\r]{2,50})'
)
RE_COMPANY_SUFFIX = re.compile(
    r'([一-鿿][一-鿿\(\)（）\w]{1,48}'
    r'(?:有限公司|有限责任公司|分公司|股份有限公司|'
    r'股份公司|贸易商行|咨询中心|工作室|'
    r'厂|店|中心|经营部|商行|社|集团|公司))'
)

# 社会信用代码匹配：先找标签格式，再找裸码
RE_CODE_LABEL = re.compile(
    r'(?:统一社会信用代码|社会信用代码|信用代码|'
    r'组织机构代码|工商注册号|注册号)[：:\s]*([A-Za-z0-9]{15,18})'
)
RE_CODE_RAW = re.compile(r'\b([A-Za-z0-9]{18})\b')
RE_CODE_15 = re.compile(r'\b([A-Za-z0-9]{15})\b')

# 注册资本匹配：先找标签格式，再找裸数字+万
RE_CAP_LABEL = re.compile(
    r'(?:注册资本|注册资金|资本总额|资本)[：:\s]*'
    r'([\d,]+(?:\.\d+)?)\s*(?:万元?|万\s*元|元)'
)
RE_CAP_RAW = re.compile(
    r'([\d,]+(?:\.\d+)?)\s*(万元|万)'
)

# 状态匹配
RE_STATUS_LABEL = re.compile(
    r'(?:状态|经营状态|当前状态|企业状态|登记状态)[：:\s]*(\S+)'
)
RE_STATUS_RAW = re.compile(r'(存续|在营|正常|开业|注销|迁出|停业|撤销)')


def _pick_best_company_name(names):
    """从多个候选名称中选最完整的一个。"""
    if not names:
        return None
    # 选最长的（通常最完整）
    return max(names, key=len).strip()


def _clean_company_name(name):
    """去除企业名称中的序号前缀。"""
    name = re.sub(
        r'^(?:(?:第[一二三四五六七八九十百\d]+[家个位]?'
        r'|\d+[.、．\)\)]'
        r'|[①②③④⑤⑥⑦⑧⑨⑩])\s*)?'
        r'(?:公司|企业|单位)?[是叫：:]?',
        '', name
    ).strip()
    return name


def _find_best_fuzzy_match(search_name, candidates, threshold=0.4):
    """从候选名称中找到与搜索名称最相似的一个。

    使用 Jaccard 字符重叠度和子串包含关系进行模糊匹配。

    Args:
        search_name: 要搜索的名称
        candidates: 候选名称列表
        threshold: 相似度阈值（0-1），低于此值视为不匹配

    Returns:
        (best_match, similarity_score) 或 (None, 0.0)
    """
    if not candidates or not search_name:
        return None, 0.0

    search_chars = set(search_name)
    best_match = None
    best_score = 0.0

    for candidate in candidates:
        # 子串包含关系 → 高置信度直接返回
        if search_name in candidate:
            return candidate, 1.0
        if candidate in search_name:
            score = len(candidate) / max(len(search_name), 1) * 0.95
            if score > best_score:
                best_score = score
                best_match = candidate
            continue

        # Jaccard 字符重叠度
        candidate_chars = set(candidate)
        intersection = len(search_chars & candidate_chars)
        union = len(search_chars | candidate_chars)
        jaccard = intersection / union if union > 0 else 0

        # 长度接近度加分（避免短词误配）
        len_ratio = min(len(search_name), len(candidate)) / max(len(search_name), len(candidate))

        # 综合评分
        score = jaccard * 0.7 + len_ratio * 0.3

        if score > best_score:
            best_score = score
            best_match = candidate

    if best_score >= threshold:
        return best_match, best_score
    return None, 0.0


def _extract_company_name(block):
    """从文本块中提取单位名称。"""
    # 方法1：标签格式
    m = RE_COMPANY_LABEL.search(block)
    if m:
        return _clean_company_name(m.group(1))
    # 方法2：后缀格式
    m = RE_COMPANY_SUFFIX.search(block)
    if m:
        return _clean_company_name(m.group(1))
    # 方法3（兜底）：文本包含信用代码或注册资本时，
    # 取行首/块首连续中文字符作为企业名
    if RE_CODE_RAW.search(block) or RE_CAP_RAW.search(block):
        m = re.match(
            r'^\s*(?:\d+[.、．\)\)]\s*)?([一-鿿]{4,30})',
            block
        )
        if m:
            return _clean_company_name(m.group(1))
    # 方法4（兜底）：查找以"公司"结尾的企业名称
    m = re.search(r'([一-鿿]{3,30}公司)', block)
    if m:
        return _clean_company_name(m.group(1))
    return None


def _extract_credit_code(block):
    """从文本块中提取社会信用代码。"""
    # 方法1：标签格式
    m = RE_CODE_LABEL.search(block)
    if m:
        return m.group(1).upper()
    # 方法2：裸18位码（过滤掉全数字的假阳性）
    for m in RE_CODE_RAW.finditer(block):
        code = m.group(1).upper()
        if not code.isdigit():
            return code
    # 方法3：15位码
    m = RE_CODE_15.search(block)
    if m:
        return m.group(1).upper()
    return ""


def _extract_capital(block):
    """从文本块中提取注册资本（万元）。"""
    # 方法1：标签格式
    m = RE_CAP_LABEL.search(block)
    if m:
        num = m.group(1).replace(",", "").replace("，", "")
        return num + "万元"
    # 方法2：裸数字+万
    m = RE_CAP_RAW.search(block)
    if m:
        num = m.group(1).replace(",", "").replace("，", "")
        return num + "万元"
    return ""


def _extract_status(block):
    """从文本块中提取企业状态。"""
    # 方法1：标签格式
    m = RE_STATUS_LABEL.search(block)
    if m:
        return m.group(1).strip()
    # 方法2：裸关键词
    m = RE_STATUS_RAW.search(block)
    if m:
        return m.group(1)
    return ""


def _split_enterprise_blocks(text):
    """
    将企业信息文本切分为单个企业的文本块。
    支持多种排版方式：
      - 空行分隔的多行块
      - 每行一个企业的单行模式
      - 连续文本块（按企业名称锚点切分）
    """
    # 统一换行符
    text = text.replace('\r\n', '\n').replace('\r', '\n')

    # ---- 策略1：按空行切分 ----
    blocks = re.split(r'\n[ \t]*\n', text)
    blocks = [b.strip() for b in blocks if b.strip()]
    if len(blocks) >= 2:
        return blocks

    # ---- 策略2：逐行切分 — 每行看是否能独立提取企业名 ----
    lines = [l.strip() for l in text.split('\n') if l.strip()]

    # 判断一行是否"像"企业开头：包含常见后缀 或 含中文字符+信用代码
    def looks_like_enterprise_start(line):
        if re.search(r'(?:有限公司|有限责任公司|分公司|股份有限公司|'
                     r'股份公司|贸易商行|'
                     r'咨询中心|经营部|商行|社|工作室|厂|店|中心|公司)', line):
            return True
        # 同时包含中文和信用代码 → 很可能是企业行
        if re.search(r'[一-鿿]', line) and re.search(r'[A-Za-z0-9]{10,}', line):
            return True
        return False

    company_starts = [i for i, l in enumerate(lines) if looks_like_enterprise_start(l)]

    if company_starts:
        # 按企业开头位置分组
        result_blocks = []
        for idx, start in enumerate(company_starts):
            end = company_starts[idx + 1] if idx + 1 < len(company_starts) else len(lines)
            chunk = '\n'.join(lines[start:end]).strip()
            if chunk:
                result_blocks.append(chunk)
        if result_blocks:
            return result_blocks

    # ---- 策略3：按编号/序号切分（1. 2. 第一家 第二家 等） ----
    numbered = re.split(
        r'(?:^|\n)\s*(?:\d+[.、．\)\)]|[①-⑩]|'
        r'第[一二三四五六七八九十百]+[家个位])',
        text
    )
    numbered = [b.strip() for b in numbered if b.strip()]
    if len(numbered) >= 2:
        return numbered

    # ---- 策略4：整体作为一块 ----
    return [text]


def _extract_enterprise_from_block(block):
    """
    从单个文本块中提取一个企业的全部字段。
    返回 dict，缺失字段为空字符串。
    """
    info = {}
    info[UNIT_NAME_COL] = _extract_company_name(block) or ""
    info[CREDIT_CODE_COL] = _extract_credit_code(block)
    info[REG_CAPITAL_COL] = _extract_capital(block)
    info[STATUS_COL] = _extract_status(block)
    info["注册资本_数值"] = parse_registered_capital(info[REG_CAPITAL_COL])
    return info


def parse_enterprise_info(filepath):
    """
    智能解析企业信息TXT文件。

    特点：
      - 不要求固定格式（制表符/逗号/自由文本均可）
      - 自动识别企业名称、社会信用代码、注册资本、状态
      - 支持单行模式、多行标签模式、空行分隔模式

    参数：
        filepath: TXT文件路径

    返回：
        {单位名称: {单位名称, 社会信用代码, 注册资本, 状态, 注册资本_数值}}
    """
    logger.info(f"开始读取企业信息TXT：{filepath}")

    if not os.path.exists(filepath):
        raise FileNotFoundError(f"企业信息文件不存在：{filepath}")

    with open(filepath, "r", encoding="utf-8-sig") as f:
        raw_text = f.read()

    if not raw_text.strip():
        logger.warning("企业信息文件为空")
        return {}

    # 切分为单个企业的文本块
    blocks = _split_enterprise_blocks(raw_text)
    logger.info(f"识别到 {len(blocks)} 个待解析的企业文本块")

    enterprise_dict = {}
    parse_warnings = []

    for idx, block in enumerate(blocks):
        info = _extract_enterprise_from_block(block)
        unit_name = info[UNIT_NAME_COL]

        if not unit_name:
            # 尝试从块内提取中文单位特征：以中文开头的第一个有意义的词组
            fallback = re.search(r'^[^\n]*?([一-鿿]{2,20})', block)
            hint = f"（附近文本：{fallback.group(1)}...）" if fallback else ""
            parse_warnings.append(f"  块{idx + 1}：未能识别单位名称{hint}")
            continue

        if unit_name in enterprise_dict:
            logger.warning(f"  单位'{unit_name}'重复出现，仅保留第一次解析结果")
            continue

        enterprise_dict[unit_name] = info

        # 记录哪些字段缺失
        missing = []
        if not info[CREDIT_CODE_COL]:
            missing.append("社会信用代码")
        if not info[REG_CAPITAL_COL]:
            missing.append("注册资本")
        if not info[STATUS_COL]:
            missing.append("状态")

        if missing:
            logger.debug(f"  '{unit_name}'：缺失字段 {', '.join(missing)}")
        else:
            logger.debug(f"  '{unit_name}'：解析完整")

    # 输出解析摘要
    logger.info(f"成功解析 {len(enterprise_dict)} 个企业信息")
    if parse_warnings:
        for w in parse_warnings:
            logger.warning(w)
    if not enterprise_dict:
        logger.warning("未能解析出任何企业信息，请检查TXT文件内容。")

    return enterprise_dict


# ============================================================
# Excel 加载与列管理
# ============================================================
def load_excel_with_style(filepath):
    """加载Excel文件，返回 (workbook, worksheet, DataFrame, 列名->列号映射)"""
    logger.info(f"读取Excel文件：{filepath}")

    if not os.path.exists(filepath):
        raise FileNotFoundError(f"文件不存在：{filepath}")

    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    # 构建列映射
    col_indices = {}
    for cell in ws[1]:
        if cell.value:
            col_indices[cell.value] = cell.column

    # 读取数据到 pandas DataFrame
    data = []
    headers = [cell.value for cell in ws[1]]

    for row in ws.iter_rows(min_row=2, values_only=True):
        data.append(row)

    df = pd.DataFrame(data, columns=headers)

    logger.info(f"成功加载表格：{ws.max_row} 行 × {ws.max_column} 列")
    return wb, ws, df, col_indices


def ensure_columns(ws, col_indices, required_columns):
    """
    确保工作表包含指定列，如果没有则创建。
    返回更新后的 {列名: 列号} 映射。
    """
    for col_name in required_columns:
        if col_name not in col_indices:
            new_col_idx = ws.max_column + 1
            cell = ws.cell(row=1, column=new_col_idx, value=col_name)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            col_indices[col_name] = new_col_idx
            logger.info(f"  添加缺失列：'{col_name}' → 第 {get_column_letter(new_col_idx)} 列")

    return col_indices


# ============================================================
# Excel 写入工具函数
# ============================================================
def append_reason(ws, row, reason_col_idx, text):
    """在指定行的'原因'列追加文本"""
    existing = ws.cell(row=row, column=reason_col_idx).value
    if existing and str(existing).strip():
        new_value = str(existing).rstrip("；;") + f"；{text}"
    else:
        new_value = text
    ws.cell(row=row, column=reason_col_idx, value=new_value)


def highlight_row(ws, row, fill, num_cols):
    """将指定行的所有单元格设为指定背景色"""
    for col in range(1, num_cols + 1):
        ws.cell(row=row, column=col).fill = fill


def find_student_name_column(headers):
    """从表头中查找学生姓名列"""
    for pattern in STUDENT_NAME_PATTERNS:
        for h in headers:
            if h and pattern in str(h):
                return h
    return None


def safe_str(value):
    """安全转换为字符串"""
    if value is None:
        return ""
    return str(value).strip()


# ============================================================
# 第一阶段：企业信息匹配与填充
# ============================================================
def phase1_match_and_fill(ws, df, enterprise_dict, col_indices):
    """
    遍历书院表格，用企业信息字典匹配并修正数据。

    处理逻辑：
      - 单位名称在企业字典中不存在 → 标红
      - 存在 → 填充/覆盖信用代码和注册资本
        - 若原值不同 → 标黄，原因列记录"信用代码错误，已纠正为：xxx"
        - 若原值相同或为空 → 静默填充

    返回 (yellow_rows, red_rows) 两个 set，其中元素为Excel行号（1-based）。
    """
    logger.info("=" * 50)
    logger.info("第一阶段：企业信息匹配与修正")
    logger.info("=" * 50)

    yellow_rows = set()
    red_rows = set()
    num_cols = ws.max_column

    reason_col_idx = col_indices.get(REASON_COL)
    credit_col_idx = col_indices.get(CREDIT_CODE_COL)
    capital_col_idx = col_indices.get(REG_CAPITAL_COL)
    unit_col_idx = col_indices.get(UNIT_NAME_COL)

    if unit_col_idx is None:
        logger.error("找不到'单位'列，无法进行匹配！")
        return yellow_rows, red_rows

    matched_count = 0
    not_found_count = 0
    corrected_count = 0

    for row_idx in range(2, ws.max_row + 1):
        unit_name = ws.cell(row=row_idx, column=unit_col_idx).value
        if not unit_name or not str(unit_name).strip():
            continue

        unit_name = str(unit_name).strip()

        if unit_name not in enterprise_dict:
            # 尝试模糊匹配（可能是书院填错了单位名称）
            fuzzy_match, score = _find_best_fuzzy_match(
                unit_name, list(enterprise_dict.keys())
            )
            if fuzzy_match:
                # 找到相似单位 → 不修改原单元格，在原因列注明正确名称
                info = enterprise_dict[fuzzy_match]

                correct_credit = info.get(CREDIT_CODE_COL, "")
                if credit_col_idx and correct_credit:
                    ws.cell(row=row_idx, column=credit_col_idx, value=correct_credit)
                correct_capital = info.get(REG_CAPITAL_COL, "")
                if capital_col_idx and correct_capital:
                    ws.cell(row=row_idx, column=capital_col_idx, value=correct_capital)

                highlight_row(ws, row_idx, YELLOW_FILL, num_cols)
                yellow_rows.add(row_idx)
                if reason_col_idx:
                    append_reason(ws, row_idx, reason_col_idx,
                                  f"单位名称有误，正确名称：{fuzzy_match}")
                corrected_count += 1
                operations_log.append(
                    f"  行{row_idx}：单位名称'{unit_name}'应改为'{fuzzy_match}'（相似度{score:.0%}），已在原因注明"
                )
                matched_count += 1
                logger.debug(f"  行{row_idx}：'{unit_name}' 模糊匹配→'{fuzzy_match}'（{score:.0%}）")
                continue
            else:
                # 企业信息未找到 → 标红
                highlight_row(ws, row_idx, RED_FILL, num_cols)
                red_rows.add(row_idx)
                not_found_count += 1
                logger.debug(f"  行{row_idx}：'{unit_name}' 未在企业信息中找到 → 标红")
                continue

        matched_count += 1
        info = enterprise_dict[unit_name]

        # ---- 处理社会信用代码 ----
        correct_credit = info.get(CREDIT_CODE_COL, "")
        if credit_col_idx and correct_credit:
            existing_val = ws.cell(row=row_idx, column=credit_col_idx).value
            existing_str = safe_str(existing_val)

            # 无论原有值是什么，都覆盖为正确值
            ws.cell(row=row_idx, column=credit_col_idx, value=correct_credit)

            # 如果原有值存在且与正确值不同，标黄并记录
            if existing_str and existing_str != correct_credit:
                # 只有当还未被标红时才操作（已标红的不再改）
                if row_idx not in red_rows:
                    highlight_row(ws, row_idx, YELLOW_FILL, num_cols)
                    yellow_rows.add(row_idx)
                if reason_col_idx:
                    append_reason(ws, row_idx, reason_col_idx,
                                  f"信用代码错误，已纠正为：{correct_credit}")
                corrected_count += 1
                operations_log.append(
                    f"  行{row_idx}：'{unit_name}' 信用代码已纠正为 {correct_credit}"
                )
                logger.debug(f"  行{row_idx}：'{unit_name}' 信用代码已纠正")

        # ---- 处理注册资本（静默覆盖） ----
        correct_capital = info.get(REG_CAPITAL_COL, "")
        if capital_col_idx and correct_capital:
            ws.cell(row=row_idx, column=capital_col_idx, value=correct_capital)

    logger.info(f"匹配结果：{matched_count} 个找到，"
                f"{not_found_count} 个未找到（已标红），"
                f"{corrected_count} 个信用代码已纠正（已标黄）")
    return yellow_rows, red_rows


# ============================================================
# 第二阶段：人数配额检查
# ============================================================
def phase2_headcount_rule(ws, df, total_df, col_indices, yellow_rows, red_rows):
    """
    人数配额检查：任何单位在总表中的总人数（已有 + 本次上报）不得超过3人。
    若超限，将该单位本次上报所有行标黄并记录原因。
    """
    logger.info("-" * 40)
    logger.info("第二阶段-1：人数配额检查")
    logger.info("-" * 40)

    reason_col_idx = col_indices.get(REASON_COL)
    unit_col_idx = col_indices.get(UNIT_NAME_COL)
    submit_col_idx = col_indices.get(THIS_SUBMIT_COL)

    if unit_col_idx is None:
        logger.error("找不到'单位'列，跳过人数配额检查")
        return yellow_rows, set()

    num_cols = ws.max_column
    new_yellows = set()
    headcount_yellow_rows = set()

    # 找到总表中的单位名称列
    total_unit_col = None
    for col in total_df.columns:
        if col and "单位" in str(col):
            total_unit_col = col
            break

    if total_unit_col is None:
        logger.error("全校总表中找不到'单位名称'列，跳过人数配额检查")
        return yellow_rows, set()

    # 收集书院表中所有单位的行分布
    unit_row_map = {}
    for row_idx in range(2, ws.max_row + 1):
        val = ws.cell(row=row_idx, column=unit_col_idx).value
        if val and str(val).strip():
            unit = str(val).strip()
            unit_row_map.setdefault(unit, []).append(row_idx)

    for unit, rows in unit_row_map.items():
        # 当前本次上报的有效行（不包含已标红的）
        active_rows = [r for r in rows if r not in red_rows]
        if not active_rows:
            continue

        # 计算本次上报人数总和（而非行数）
        current_count = 0
        for r_idx in active_rows:
            submit_val = ws.cell(row=r_idx, column=submit_col_idx).value if submit_col_idx else None
            try:
                current_count += int(float(submit_val))
            except (ValueError, TypeError):
                current_count += 1  # 无法转换则算1人

        # 总表中该单位已有人数
        # 优先使用 全校已有人数 列（用于支持总表去重后的行数统计）
        try:
            unit_rows = total_df[total_df[total_unit_col].astype(str).str.strip() == unit]
            if unit_rows.empty:
                existing_count = 0
            else:
                # 查找是否包含"全校已有人数"列
                total_count_col_name = None
                for col in total_df.columns:
                    if col and TOTAL_COUNT_COL in str(col):
                        total_count_col_name = col
                        break
                if total_count_col_name:
                    existing_count = int(pd.to_numeric(
                        unit_rows[total_count_col_name], errors='coerce'
                    ).sum())
                else:
                    existing_count = len(unit_rows)
        except Exception:
            existing_count = 0

        total_count = existing_count + current_count

        if total_count > MAX_PEOPLE_PER_UNIT:
            y = MAX_PEOPLE_PER_UNIT - existing_count  # 最多还能增加的人数
            reason = f"全校已有{existing_count}人，最多还能增加{y}人"

            # 从前往后累计，找到在名额内能容纳的行
            cumulative = 0
            split_idx = 0
            for r_idx in active_rows:
                if cumulative >= y:
                    break
                submit_val = ws.cell(row=r_idx, column=submit_col_idx).value if submit_col_idx else None
                try:
                    people = int(float(submit_val))
                except (ValueError, TypeError):
                    people = 1
                if cumulative + people > y:
                    break
                cumulative += people
                split_idx += 1

            # split_idx 之后的行都是超额的
            excess_rows = active_rows[split_idx:]
            excess_people = current_count - cumulative
            for r_idx in excess_rows:
                if r_idx not in red_rows:
                    highlight_row(ws, r_idx, YELLOW_FILL, num_cols)
                    headcount_yellow_rows.add(r_idx)
                    new_yellows.add(r_idx)
                    if reason_col_idx:
                        append_reason(ws, r_idx, reason_col_idx, reason)
            operations_log.append(
                f"  '{unit}'：全校已有{existing_count}人，本次上报{current_count}人，"
                f"最多增{y}人，{len(excess_rows)}行（共{excess_people}人）因超额标黄"
            )
            logger.info(f"  '{unit}'：全校已有{existing_count}人，"
                        f"本次上报{current_count}人，最多增{y}人 → {len(excess_rows)}行超出")

    yellow_rows.update(new_yellows)
    logger.info(f"人数配额检查完成，新增标黄 {len(new_yellows)} 行"
                f"（其中 {len(headcount_yellow_rows)} 行因人数超额）")
    return yellow_rows, headcount_yellow_rows


# ============================================================
# 第二阶段-2：注册资本审核
# ============================================================
def phase2_capital_rule(ws, df, enterprise_dict, col_indices, yellow_rows, red_rows):
    """
    注册资本审核规则：
      - 20 <= 注册资本 < 50：该单位最多上报2人，超过则标黄
      - 10 < 注册资本 <= 20：该单位最多上报2人，超过则标黄
      - 1 <= 注册资本 <= 10：该单位最多上报1人，超过则标黄
      - 0 < 注册资本 < 1：标黄，需老师确认
      - 注册资本为空/0（如分公司）：跳过检查
    """
    logger.info("-" * 40)
    logger.info("第二阶段-2：注册资本审核")
    logger.info("-" * 40)

    reason_col_idx = col_indices.get(REASON_COL)
    unit_col_idx = col_indices.get(UNIT_NAME_COL)
    submit_col_idx = col_indices.get(THIS_SUBMIT_COL)
    num_cols = ws.max_column
    new_yellows = set()

    def _sum_submit(rows):
        """计算 rows 中 本次上报人数 的总和"""
        total = 0
        for r in rows:
            v = ws.cell(row=r, column=submit_col_idx).value if submit_col_idx else None
            try:
                total += int(float(v))
            except (ValueError, TypeError):
                total += 1
        return total

    if unit_col_idx is None:
        logger.error("找不到'单位'列，跳过注册资本审核")
        return yellow_rows

    # 按单位分组
    unit_rows = {}
    for row_idx in range(2, ws.max_row + 1):
        val = ws.cell(row=row_idx, column=unit_col_idx).value
        if val and str(val).strip():
            unit = str(val).strip()
            unit_rows.setdefault(unit, []).append(row_idx)

    for unit, rows in unit_rows.items():
        if unit not in enterprise_dict:
            continue

        cap_value = enterprise_dict[unit].get("注册资本_数值", 0.0)

        # 注册资本为空或0（如分公司）→ 跳过
        if cap_value == 0.0:
            continue

        # 有效行（不包含已标红的）
        active_rows = [r for r in rows if r not in red_rows]
        if not active_rows:
            continue

        active_count = _sum_submit(active_rows)

        if 20 <= cap_value < 50:
            if active_count > 2:
                reason = "注册资本不足50万，最多报2人"
                for r_idx in active_rows:
                    highlight_row(ws, r_idx, YELLOW_FILL, num_cols)
                    new_yellows.add(r_idx)
                    if reason_col_idx:
                        append_reason(ws, r_idx, reason_col_idx, reason)
                operations_log.append(
                    f"  '{unit}'：注册资本{cap_value}万（20~50万区间），"
                    f"上报{active_count}人超过2人限制 → 标黄"
                )
                logger.info(f"  '{unit}'：注册资本{cap_value}万元"
                            f"（20~50万区间），上报{active_count}人 > 2人 → 标黄")
        elif 0 < cap_value < 1:
            reason = "注册资本少于1万，需反馈老师确认"
            for r_idx in active_rows:
                highlight_row(ws, r_idx, YELLOW_FILL, num_cols)
                new_yellows.add(r_idx)
                if reason_col_idx:
                    append_reason(ws, r_idx, reason_col_idx, reason)
            operations_log.append(
                f"  '{unit}'：注册资本{cap_value}万（少于1万）→ 标黄，需反馈老师确认"
            )
            logger.info(f"  '{unit}'：注册资本{cap_value}万元"
                        f"（少于1万）→ 标黄")
        elif 1 <= cap_value <= 10:
            if active_count > 1:
                reason = "注册资本不超过10万，最多报1人"
                for r_idx in active_rows:
                    highlight_row(ws, r_idx, YELLOW_FILL, num_cols)
                    new_yellows.add(r_idx)
                    if reason_col_idx:
                        append_reason(ws, r_idx, reason_col_idx, reason)
                operations_log.append(
                    f"  '{unit}'：注册资本{cap_value}万（1~10万区间），"
                    f"上报{active_count}人超过1人限制 → 标黄"
                )
                logger.info(f"  '{unit}'：注册资本{cap_value}万元"
                            f"（1~10万区间），上报{active_count}人 > 1人 → 标黄")
        elif 10 < cap_value <= 20:
            if active_count > 2:
                reason = "注册资本不超过20万，最多报2人"
                for r_idx in active_rows:
                    highlight_row(ws, r_idx, YELLOW_FILL, num_cols)
                    new_yellows.add(r_idx)
                    if reason_col_idx:
                        append_reason(ws, r_idx, reason_col_idx, reason)
                operations_log.append(
                    f"  '{unit}'：注册资本{cap_value}万（10~20万区间），"
                    f"上报{active_count}人超过2人限制 → 标黄"
                )
                logger.info(f"  '{unit}'：注册资本{cap_value}万元"
                            f"（10~20万区间），上报{active_count}人 > 2人 → 标黄")

    yellow_rows.update(new_yellows)
    logger.info(f"注册资本审核完成，新增标黄 {len(new_yellows)} 行")
    return yellow_rows


# ============================================================
# 第二阶段-3：其他规则（个体工商户等）
# ============================================================
def phase2_other_rules(ws, df, col_indices, yellow_rows, red_rows):
    """
    其他规则审核：
      - 社会信用代码以"92"开头 → 个体工商户，标黄
    """
    logger.info("-" * 40)
    logger.info("第二阶段-3：其他规则审核")
    logger.info("-" * 40)

    reason_col_idx = col_indices.get(REASON_COL)
    credit_col_idx = col_indices.get(CREDIT_CODE_COL)
    num_cols = ws.max_column
    new_yellows = set()

    if credit_col_idx is None:
        logger.error("找不到'社会信用代码'列，跳过个体工商户审核")
        return yellow_rows

    for row_idx in range(2, ws.max_row + 1):
        if row_idx in red_rows:
            continue

        credit_code = ws.cell(row=row_idx, column=credit_col_idx).value
        if credit_code and str(credit_code).strip().startswith("92"):
            highlight_row(ws, row_idx, YELLOW_FILL, num_cols)
            new_yellows.add(row_idx)
            if reason_col_idx:
                append_reason(ws, row_idx, reason_col_idx, "个体工商户不能上报")

    yellow_rows.update(new_yellows)
    if new_yellows:
        operations_log.append(f"  个体工商户检查：{len(new_yellows)} 行因92开头被标黄")
        logger.info(f"个体工商户检查完成，{len(new_yellows)} 行被标黄（不可合并）")
    else:
        logger.info("个体工商户检查完成，无匹配项")
    return yellow_rows, new_yellows


# ============================================================
# 第二阶段-4：企业状态审核
# ============================================================
def phase2_status_rule(ws, df, enterprise_dict, col_indices, yellow_rows, red_rows):
    """
    企业状态审核：天眼查上企业状态为"注销"的不能上报。
    从企业信息字典中获取状态字段，若为"注销"则标黄。
    """
    logger.info("-" * 40)
    logger.info("第二阶段-4：企业状态审核")
    logger.info("-" * 40)

    reason_col_idx = col_indices.get(REASON_COL)
    credit_col_idx = col_indices.get(CREDIT_CODE_COL)
    unit_col_idx = col_indices.get(UNIT_NAME_COL)
    num_cols = ws.max_column
    new_yellows = set()

    if unit_col_idx is None:
        logger.error("找不到'单位'列，跳过企业状态审核")
        return yellow_rows

    for row_idx in range(2, ws.max_row + 1):
        if row_idx in red_rows:
            continue

        unit_name = ws.cell(row=row_idx, column=unit_col_idx).value
        if not unit_name or not str(unit_name).strip():
            continue

        unit_name = str(unit_name).strip()
        if unit_name not in enterprise_dict:
            continue

        # 从企业信息字典获取状态
        status = enterprise_dict[unit_name].get(STATUS_COL, "")
        if status == "注销":
            highlight_row(ws, row_idx, YELLOW_FILL, num_cols)
            new_yellows.add(row_idx)
            if reason_col_idx:
                append_reason(ws, row_idx, reason_col_idx,
                              '企业状态为"注销"，不能上报')

    yellow_rows.update(new_yellows)
    if new_yellows:
        operations_log.append(f'  企业状态审核：{len(new_yellows)} 行因状态为"注销"被标黄（不可合并）')
        logger.info(f"企业状态审核完成，{len(new_yellows)} 行被标黄（状态为注销，不可合并）")
    else:
        logger.info("企业状态审核完成，无注销状态企业")
    return yellow_rows, new_yellows


# ============================================================
# 第三阶段：数据合并到总表（核心难点）
# ============================================================
def phase3_merge(ws, total_filepath, enterprise_dict, col_indices,
                 yellow_rows, red_rows, headcount_yellow_rows=None,
                 unmergeable_yellows=None):
    """
    将书院表格中审核通过的行合并到全校总表。

    合并逻辑：
      1. 跳过标红行、个体工商户和注销状态标黄的行
      2. 其他标黄的行（人数超限额、信用代码、注册资本等）仍可合并，使用 y 值
      3. 只复制 单位名称、社会信用代码、注册资本、本次上报人数 四列
      4. 本次上报人数 → 全校已有人数（标黄行用 y 值）
      5. 合并后按单位名称去重，累加全校已有人数

    返回更新后的总表 workbook 对象。
    """
    logger.info("=" * 50)
    logger.info("第三阶段：合并数据到全校总表")
    logger.info("=" * 50)

    if headcount_yellow_rows is None:
        headcount_yellow_rows = set()
    if unmergeable_yellows is None:
        unmergeable_yellows = set()

    unit_col_idx = col_indices.get(UNIT_NAME_COL)
    credit_col_idx = col_indices.get(CREDIT_CODE_COL)
    capital_col_idx = col_indices.get(REG_CAPITAL_COL)
    reason_col_idx = col_indices.get(REASON_COL)
    submit_col_idx = col_indices.get(THIS_SUBMIT_COL)

    if unit_col_idx is None:
        logger.error("找不到'单位'列，无法合并")
        return None

    # ---- 加载全校总表 ----
    if not os.path.exists(total_filepath):
        logger.error(f"全校总表文件不存在：{total_filepath}")
        return None

    total_wb = openpyxl.load_workbook(total_filepath)
    total_ws = total_wb.active

    # 构建总表列名 → 列号映射
    total_col_indices = {}
    for cell in total_ws[1]:
        if cell.value:
            total_col_indices[cell.value] = cell.column

    # 确保总表有"全校已有人数"列
    total_col_indices = ensure_columns(total_ws, total_col_indices,
                                        [TOTAL_COUNT_COL])

    # 查找各列在总表中的位置
    total_unit_col = None
    total_credit_col = None
    total_capital_col = None
    total_count_col = None
    for hdr, idx in total_col_indices.items():
        if hdr and "单位" in str(hdr):
            total_unit_col = idx
        if hdr == CREDIT_CODE_COL:
            total_credit_col = idx
        if hdr == REG_CAPITAL_COL:
            total_capital_col = idx
        if hdr == TOTAL_COUNT_COL:
            total_count_col = idx

    if total_unit_col is None:
        logger.error("全校总表中找不到'单位'列，无法合并")
        total_wb.close()
        return None

    # ---- 统计总表各单位的累计已有人数 ----
    last_total_row = total_ws.max_row
    existing_counts = {}
    for row_idx in range(2, last_total_row + 1):
        unit_val = total_ws.cell(row=row_idx, column=total_unit_col).value
        if unit_val and str(unit_val).strip():
            unit = str(unit_val).strip()
            count_val = 0
            if total_count_col is not None:
                cell_v = total_ws.cell(row=row_idx, column=total_count_col).value
                try:
                    count_val = int(float(cell_v)) if cell_v is not None else 0
                except (ValueError, TypeError):
                    count_val = 0
            else:
                count_val = 1  # 没有全校已有人数列则每行算1人
            existing_counts[unit] = existing_counts.get(unit, 0) + count_val

    # ---- 收集书院表中各单位的已审核通过的行及其本次上报人数 ----
    unit_approved = {}  # {unit: [(row_idx, people), ...]}

    for row_idx in range(2, ws.max_row + 1):
        if row_idx in red_rows or row_idx in unmergeable_yellows:
            continue
        val = ws.cell(row=row_idx, column=unit_col_idx).value
        if val and str(val).strip():
            unit = str(val).strip()
            submit_val = ws.cell(row=row_idx, column=submit_col_idx).value if submit_col_idx else None
            try:
                people = int(float(submit_val))
            except (ValueError, TypeError):
                people = 1
            unit_approved.setdefault(unit, []).append((row_idx, people))

    logger.info(f"待合并单位数：{len(unit_approved)}")

    # ---- 列映射定义 ----
    # 确定要复制的列对：(书院列名, 总表列名, 是否用企业字典修正值)
    column_mappings = []
    if unit_col_idx and total_unit_col:
        column_mappings.append((UNIT_NAME_COL, None, False))
    if credit_col_idx and total_credit_col:
        column_mappings.append((CREDIT_CODE_COL, None, True))
    if capital_col_idx and total_capital_col:
        column_mappings.append((REG_CAPITAL_COL, None, True))

    # ---- 执行合并 ----
    merged_total = 0
    skipped_total = 0

    for unit, approved_rows in unit_approved.items():
        existing = existing_counts.get(unit, 0)
        can_add = MAX_PEOPLE_PER_UNIT - existing
        y = max(0, can_add)  # 最多还能增加的人数

        # 计算本次上报人数总和
        submit_total = sum(p for _, p in approved_rows)

        if can_add <= 0 or submit_total <= 0:
            # 名额已满，全部跳过
            for r_idx, _ in approved_rows:
                if reason_col_idx:
                    append_reason(ws, r_idx, reason_col_idx,
                                  "因全校该单位名额已满，未合并")
                highlight_row(ws, r_idx, YELLOW_FILL, ws.max_column)
                yellow_rows.add(r_idx)
            skipped_total += len(approved_rows)
            operations_log.append(
                f"  '{unit}'：名额已满（总表已有{existing}人），跳过{len(approved_rows)}行（共{submit_total}人）"
            )
            logger.info(f"  '{unit}'：名额已满（总表已有{existing}人），"
                        f"跳过{len(approved_rows)}行（共{submit_total}人）")
            continue

        # 从前往后取行，直到名额用完
        # 标黄行合并时使用 y 值（写入总表的值）而非原人数
        merge_rows = []
        skip_rows = []
        accumulated = 0
        for item in approved_rows:
            r_idx, people = item
            effective_people = y if r_idx in yellow_rows else people
            if accumulated < can_add and accumulated + effective_people <= can_add:
                merge_rows.append(item)
                accumulated += effective_people
            else:
                skip_rows.append(item)

        # --- 执行合并：逐行追加到总表 ---
        for r_idx, _ in merge_rows:
            new_row_idx = last_total_row + 1
            last_total_row = new_row_idx

            # 复制 单位名称
            total_ws.cell(row=new_row_idx, column=total_unit_col,
                          value=ws.cell(row=r_idx, column=unit_col_idx).value)

            # 复制 社会信用代码（使用企业字典修正值）
            if total_credit_col is not None:
                val = enterprise_dict.get(unit, {}).get(CREDIT_CODE_COL, "")
                if not val:
                    val = ws.cell(row=r_idx, column=credit_col_idx).value if credit_col_idx else ""
                total_ws.cell(row=new_row_idx, column=total_credit_col, value=val)

            # 复制 注册资本（使用企业字典修正值）
            if total_capital_col is not None:
                val = enterprise_dict.get(unit, {}).get(REG_CAPITAL_COL, "")
                if not val:
                    val = ws.cell(row=r_idx, column=capital_col_idx).value if capital_col_idx else ""
                total_ws.cell(row=new_row_idx, column=total_capital_col, value=val)

            # 本次上报人数 → 全校已有人数
            if total_count_col is not None:
                if r_idx in yellow_rows:
                    # 标黄行：使用 y 值（剩余可增人数）
                    total_ws.cell(row=new_row_idx, column=total_count_col, value=y)
                else:
                    # 非标黄行：使用本次上报人数
                    submit_val = ws.cell(row=r_idx, column=submit_col_idx).value if submit_col_idx else 1
                    try:
                        submit_val = int(float(submit_val))
                    except (ValueError, TypeError):
                        submit_val = 1
                    total_ws.cell(row=new_row_idx, column=total_count_col, value=submit_val)

            merged_total += 1

        # --- 处理超出行（未能合并的行） ---
        for r_idx, _ in skip_rows:
            if reason_col_idx:
                append_reason(ws, r_idx, reason_col_idx,
                              "因全校该单位名额不足，未合并")
            highlight_row(ws, r_idx, YELLOW_FILL, ws.max_column)
            yellow_rows.add(r_idx)
            skipped_total += 1

        if skip_rows:
            operations_log.append(
                f"  '{unit}'：总表已有{existing}人，合并{len(merge_rows)}行，跳过{len(skip_rows)}行"
            )
            logger.info(f"  '{unit}'：总表已有{existing}人，"
                        f"合并{len(merge_rows)}行，跳过{len(skip_rows)}行")

    logger.info(f"合并完成：成功合并 {merged_total} 行，跳过 {skipped_total} 行")
    operations_log.append(f"  合并总表：成功合并 {merged_total} 行，跳过 {skipped_total} 行")

    # ============================================================
    # 去重：合并后按单位名称去重，累加 全校已有人数
    # ============================================================
    if total_count_col is not None and merged_total > 0:
        logger.info("合并后去重：按单位名称去重，累加全校已有人数...")

        # 读取总表所有数据
        total_headers = []
        for cell in total_ws[1]:
            total_headers.append(cell.value)

        total_data_rows = []
        for row in total_ws.iter_rows(min_row=2, values_only=True):
            total_data_rows.append(row)

        if total_data_rows:
            # 按单位名称分组
            unit_data = {}  # {unit: (row_data_list)}
            unit_order = []  # 保持顺序
            for row_data in total_data_rows:
                # 找到单位名称列索引
                unit_val = None
                for i, h in enumerate(total_headers):
                    if h and total_unit_col == i + 1:
                        # Column index is 1-based, list index is 0-based
                        pass
                # Simpler: use total_unit_col (1-based) directly
                unit_idx_0 = total_unit_col - 1  # 0-based
                if unit_idx_0 < len(row_data):
                    unit_val = row_data[unit_idx_0]

                if unit_val and str(unit_val).strip():
                    u = str(unit_val).strip()
                    if u not in unit_data:
                        unit_data[u] = []
                        unit_order.append(u)
                    unit_data[u].append(row_data)

            # 去重：每单位保留一行，累加全校已有人数
            count_idx_0 = total_count_col - 1 if total_count_col else -1
            deduped_rows = []
            dup_merged_count = 0
            for u in unit_order:
                rows = unit_data[u]
                if len(rows) == 1:
                    deduped_rows.append(rows[0])
                else:
                    # 多行 → 保留第一行，但累加全校已有人数
                    merged_row = list(rows[0])
                    if count_idx_0 >= 0:
                        total_sum = 0
                        for r in rows:
                            try:
                                total_sum += float(r[count_idx_0]) if r[count_idx_0] is not None else 0
                            except (ValueError, TypeError):
                                total_sum += 0
                        merged_row[count_idx_0] = total_sum
                    deduped_rows.append(merged_row)
                    dup_merged_count += len(rows) - 1

            if dup_merged_count > 0:
                logger.info(f"  去重完成：合并了 {dup_merged_count} 个重复单位行")
                operations_log.append(f"  总表去重：合并了 {dup_merged_count} 个重复单位行")

                # 清空数据行并重写
                total_ws.delete_rows(2, total_ws.max_row)
                for row_data in deduped_rows:
                    total_ws.append(row_data)
                logger.info(f"  去重后总表共 {len(deduped_rows)} 行")

    return total_wb


# ============================================================
# 第四阶段：输出结果
# ============================================================
def phase4_output(ws, total_wb, col_indices, yellow_rows, red_rows, args):
    """
    输出三个文件：
      1. 书院待筛选表格_已处理_YYYYMMDD.xlsx  （处理后的书院表，含颜色标注）
      2. 全校数据总表格_更新后_YYYYMMDD.xlsx   （合并后的总表）
      3. 出错学生信息_YYYYMMDD.xlsx            （仅当检测到学生姓名列且存在标黄行时生成）
    """
    logger.info("=" * 50)
    logger.info("第四阶段：输出处理结果")
    logger.info("=" * 50)

    today = datetime.now().strftime("%Y%m%d_%H%M")
    output_dir = "output"

    # ========== 输出文件1：处理后的书院表格 ==========
    output1_path = os.path.join(output_dir, f"书院待筛选表格_已处理_{today}.xlsx")
    logger.info(f"正在保存处理后书院表格...")
    try:
        # 将数值列中的文本格式数字转为实际数值
        numeric_col_names = [THIS_SUBMIT_COL, TOTAL_COUNT_COL]
        for col_name in numeric_col_names:
            col_idx = col_indices.get(col_name)
            if col_idx:
                for row in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row, column=col_idx)
                    if cell.value is not None and isinstance(cell.value, str):
                        try:
                            cell.value = int(float(cell.value))
                        except (ValueError, TypeError):
                            pass  # 保留原文本（如注册资本"200万"）
        # 设置 header 样式（让表头更清晰）
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL

        # 设置所有数据行文字垂直居中
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row,
                                max_col=ws.max_column):
            for cell in row:
                cell.alignment = Alignment(vertical="center", wrap_text=True)

        ws.parent.save(output1_path)
        logger.info(f"  ✓ 输出文件1：{output1_path}")
        operations_log.append(f"  输出文件1：处理后的书院表格 → {output1_path}")
    except Exception as e:
        logger.error(f"  ✗ 保存书院表格失败：{e}")

    # ========== 输出文件2：更新后的全校总表 ==========
    if total_wb is not None:
        total_ws_out = total_wb.active
        # 将全校已有人数列的文本数字转为数值
        total_count_col_idx = None
        for cell in total_ws_out[1]:
            if cell.value == TOTAL_COUNT_COL:
                total_count_col_idx = cell.column
                break
        if total_count_col_idx:
            for row in range(2, total_ws_out.max_row + 1):
                cell = total_ws_out.cell(row=row, column=total_count_col_idx)
                if cell.value is not None and isinstance(cell.value, str):
                    try:
                        cell.value = int(float(cell.value))
                    except (ValueError, TypeError):
                        pass

        output2_path = os.path.join(output_dir, f"全校数据总表格_更新后_{today}.xlsx")
        logger.info(f"正在保存更新后全校总表...")
        try:
            total_wb.save(output2_path)
            logger.info(f"  ✓ 输出文件2：{output2_path}")
            operations_log.append(f"  输出文件2：更新后总表 → {output2_path}")
        except Exception as e:
            logger.error(f"  ✗ 保存全校总表失败：{e}")
        finally:
            total_wb.close()
    else:
        logger.warning("  全校总表未更新，跳过输出文件2")

    # ========== 输出文件3：出错学生信息表（条件触发） ==========
    headers = [cell.value for cell in ws[1]]
    student_col = find_student_name_column(headers)

    if student_col:
        student_col_idx = col_indices.get(student_col)
        unit_col_idx = col_indices.get(UNIT_NAME_COL)
        reason_col_idx = col_indices.get(REASON_COL)

        # 收集所有标黄行（去重按行号）
        error_rows = sorted(yellow_rows - red_rows) if yellow_rows else []

        if error_rows:
            # 创建新工作簿
            error_wb = openpyxl.Workbook()
            error_ws = error_wb.active
            error_ws.title = "出错学生信息"

            # 写入表头
            err_headers = [UNIT_NAME_COL, student_col, REASON_COL]
            for col_idx, h in enumerate(err_headers, start=1):
                cell = error_ws.cell(row=1, column=col_idx, value=h)
                cell.font = HEADER_FONT
                cell.fill = PatternFill(start_color="FCE4D6",
                                        end_color="FCE4D6",
                                        fill_type="solid")

            # 写入数据行
            row_num = 2
            for ws_row in error_rows:
                u_val = ws.cell(row=ws_row, column=unit_col_idx).value if unit_col_idx else ""
                s_val = ws.cell(row=ws_row, column=student_col_idx).value if student_col_idx else ""
                r_val = ws.cell(row=ws_row, column=reason_col_idx).value if reason_col_idx else ""

                error_ws.cell(row=row_num, column=1, value=u_val)
                error_ws.cell(row=row_num, column=2, value=s_val)
                error_ws.cell(row=row_num, column=3, value=r_val)
                row_num += 1

            output3_path = os.path.join(output_dir, f"出错学生信息_{today}.xlsx")
            error_wb.save(output3_path)
            error_wb.close()
            logger.info(f"  ✓ 输出文件3：{output3_path}（共 {len(error_rows)} 行）")
            operations_log.append(f"  输出文件3：出错学生信息 → {output3_path}（{len(error_rows)} 行）")
        else:
            logger.info("  - 无标黄行，跳过生成出错学生信息表")
    else:
        logger.info("  - 未检测到学生姓名列，跳过生成出错学生信息表")


# ============================================================
# 汇总统计
# ============================================================
def print_summary(ws, yellow_rows, red_rows):
    """打印处理汇总和详细操作日志"""
    total_data_rows = ws.max_row - 1
    yellow_count = len(yellow_rows)
    red_count = len(red_rows)
    clean_count = total_data_rows - yellow_count - red_count

    logger.info("")
    logger.info("=" * 60)
    logger.info("  处理完成 — 汇总报告")
    logger.info("=" * 60)
    logger.info(f"  总数据行数：     {total_data_rows}")
    logger.info(f"  审核通过：       {clean_count} 行")
    logger.info(f"  标黄（警告）：   {yellow_count} 行")
    logger.info(f"  标红（未找到）： {red_count} 行")
    logger.info("=" * 60)

    # ---- 输出详细操作日志 ----
    if operations_log:
        logger.info("")
        logger.info("-" * 60)
        logger.info("  本次操作明细")
        logger.info("-" * 60)
        for op in operations_log:
            logger.info(op)
        logger.info("-" * 60)
        logger.info(f"  共 {len(operations_log)} 项操作")
        logger.info("=" * 60)


# ============================================================
# 主函数
# ============================================================
def main():
    """主流程入口"""
    args = parse_args()

    # 确保 input/ 和 output/ 目录存在
    os.makedirs("input", exist_ok=True)
    os.makedirs("output", exist_ok=True)

    logger.info("")
    logger.info("╔" + "═" * 58 + "╗")
    logger.info("║           就业数据自动化处理工具              ║")
    logger.info("╚" + "═" * 58 + "╝")
    logger.info("输入文件：")
    logger.info(f"  ① 新数据表格：{args.new}")
    logger.info(f"  ② 全校总表：  {args.total}")
    logger.info(f"  ③ 企业信息：  {args.enterprise}")
    logger.info("输出目录：output/")
    logger.info("")

    # 清空操作日志
    global operations_log
    operations_log = []

    try:
        # ========== 解析企业信息 ==========
        enterprise_dict = parse_enterprise_info(args.enterprise)

        # ========== 加载书院待筛选表格 ==========
        wb, ws, df, col_indices = load_excel_with_style(args.new)

        # 确保必要列存在
        col_indices = ensure_columns(
            ws, col_indices,
            [REASON_COL, CREDIT_CODE_COL, REG_CAPITAL_COL, THIS_SUBMIT_COL]
        )

        # ========== 第一阶段：匹配与填充 ==========
        yellow_rows, red_rows = phase1_match_and_fill(
            ws, df, enterprise_dict, col_indices
        )

        # ========== 加载全校总表（用于规则审核） ==========
        logger.info("")
        logger.info(f"加载全校总表用于规则审核：{args.total}")
        if not os.path.exists(args.total):
            raise FileNotFoundError(f"全校总表文件不存在：{args.total}")

        total_wb_for_check = openpyxl.load_workbook(args.total)
        total_ws_check = total_wb_for_check.active
        total_data = []
        total_headers = [cell.value for cell in total_ws_check[1]]
        for row in total_ws_check.iter_rows(min_row=2, values_only=True):
            total_data.append(row)
        total_df = pd.DataFrame(total_data, columns=total_headers)
        total_wb_for_check.close()

        # 记录总表已有数据概况
        if total_df is not None and not total_df.empty:
            operations_log.append(
                f"  全校总表加载：{len(total_df)} 行数据"
            )

        # ========== 准备工作：新增全校已有人数列并填充 x 值 ==========
        logger.info("")
        logger.info("准备工作：填充全校已有人数列")
        # 计算各单位在总表中的累计已有人数
        total_unit_col_name = None
        for col in total_df.columns:
            if col and "单位" in str(col):
                total_unit_col_name = col
                break

        total_count_col_name = None
        for col in total_df.columns:
            if col and TOTAL_COUNT_COL in str(col):
                total_count_col_name = col
                break

        existing_x_map = {}  # {unit: x}
        if total_unit_col_name is not None:
            for _, row in total_df.iterrows():
                unit = str(row[total_unit_col_name]).strip() if pd.notna(row[total_unit_col_name]) else ""
                if unit:
                    count_val = 0
                    if total_count_col_name and pd.notna(row.get(total_count_col_name)):
                        conv = pd.to_numeric(row[total_count_col_name], errors='coerce')
                        count_val = int(conv) if pd.notna(conv) else 0
                    else:
                        count_val = 1  # 没有全校已有人数列则每行算1人
                    existing_x_map[unit] = existing_x_map.get(unit, 0) + count_val

        # 在书院表中新增/填充"全校已有人数"列
        col_indices = ensure_columns(ws, col_indices, [TOTAL_COUNT_COL])
        total_count_col_idx = col_indices.get(TOTAL_COUNT_COL)
        unit_col_idx = col_indices.get(UNIT_NAME_COL)
        prep_filled = 0
        for row_idx in range(2, ws.max_row + 1):
            unit_val = ws.cell(row=row_idx, column=unit_col_idx).value if unit_col_idx else None
            if unit_val and str(unit_val).strip():
                unit_name = str(unit_val).strip()
                x_val = existing_x_map.get(unit_name, 0)
                ws.cell(row=row_idx, column=total_count_col_idx, value=x_val)
                prep_filled += 1
        operations_log.append(
            f"  全校已有人数准备：已填充 {prep_filled} 行"
        )
        logger.info(f"准备工作完成，已填充 {prep_filled} 行的'全校已有人数'值")

        # ========== 第二阶段：业务规则审核 ==========
        yellow_rows, headcount_yellow_rows = phase2_headcount_rule(
            ws, df, total_df, col_indices, yellow_rows, red_rows
        )
        yellow_rows = phase2_capital_rule(
            ws, df, enterprise_dict, col_indices, yellow_rows, red_rows
        )
        yellow_rows, biz_yellow_rows = phase2_other_rules(
            ws, df, col_indices, yellow_rows, red_rows
        )
        yellow_rows, status_yellow_rows = phase2_status_rule(
            ws, df, enterprise_dict, col_indices, yellow_rows, red_rows
        )

        # 收集不可合并行
        unmergeable_yellows = biz_yellow_rows | status_yellow_rows

        # ========== 第三阶段：合并到总表 ==========
        total_wb = phase3_merge(
            ws, args.total, enterprise_dict, col_indices,
            yellow_rows, red_rows, headcount_yellow_rows, unmergeable_yellows
        )

        # ========== 第四阶段：输出 ==========
        phase4_output(ws, total_wb, col_indices, yellow_rows, red_rows, args)

        # ========== 打印汇总 ==========
        print_summary(ws, yellow_rows, red_rows)

        return 0

    except FileNotFoundError as e:
        logger.error(f"")
        logger.error(f"错误：找不到文件 — {e}")
        logger.error(f"请确保输入文件路径正确，或使用 --new/--total/--enterprise 参数指定。")
        return 1
    except PermissionError as e:
        logger.error(f"")
        logger.error(f"错误：权限不足 — {e}")
        logger.error(f"请确保文件未被其他程序打开。")
        return 1
    except Exception as e:
        logger.error(f"")
        logger.error(f"处理过程中出现未预期错误：")
        logger.error(f"  {e}")
        import traceback
        traceback.print_exc()
        return 1


if __name__ == "__main__":
    sys.exit(main())
